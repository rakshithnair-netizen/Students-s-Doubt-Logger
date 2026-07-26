import os
import json
import sqlite3

class DoubtStore:
    def __init__(self):
        # Resolve database paths relative to this file's directory
        self.base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.config_path = os.path.join(self.base_dir, "config.json")
        self.env_path = os.path.join(self.base_dir, ".env")
        
        # Paths to old Excel files (for migration)
        self.users_path = os.path.join(self.base_dir, "users.xlsx")
        self.doubts_path = os.path.join(self.base_dir, "doubts.xlsx")
        self.subjects_path = os.path.join(self.base_dir, "subjects.xlsx")

        self.default_users = []
        self.default_subjects = []

        # Load environment configurations
        self._load_env()
        self._load_config()

        # Connect to SQLite database
        db_file = os.environ.get("DB_PATH") or self.config.get("db_path")
        if not db_file:
            raise KeyError("Configuration key 'db_path' / DB_PATH is missing from config.json and environment")
        self.db_path = os.path.join(self.base_dir, db_file)
        
        self.conn = sqlite3.connect(self.db_path)
        self._init_db()

    def _load_env(self):
        if os.path.exists(self.env_path):
            try:
                with open(self.env_path, "r") as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith("#") and "=" in line:
                            key, val = line.split("=", 1)
                            os.environ[key.strip()] = val.strip()
            except Exception:
                pass

    def _load_config(self):
        self.config = {}
        if os.path.exists(self.config_path):
            try:
                with open(self.config_path, "r") as f:
                    self.config = json.load(f)
            except Exception:
                pass

        env_users = os.environ.get("DEFAULT_USERS")
        if env_users:
            try:
                self.default_users = json.loads(env_users)
            except Exception:
                self.default_users = []
        else:
            self.default_users = self.config.get("default_users", [])

        env_subjects = os.environ.get("DEFAULT_SUBJECTS")
        if env_subjects:
            try:
                self.default_subjects = json.loads(env_subjects)
            except Exception:
                self.default_subjects = []
        else:
            self.default_subjects = self.config.get("default_subjects", [])

    def _init_db(self):
        cursor = self.conn.cursor()
        
        # Create tables
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                username TEXT PRIMARY KEY,
                password TEXT NOT NULL,
                role TEXT NOT NULL
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS subjects (
                subject TEXT PRIMARY KEY
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS doubts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student TEXT NOT NULL,
                subject TEXT NOT NULL,
                teacher TEXT NOT NULL,
                severity TEXT NOT NULL,
                urgent TEXT NOT NULL,
                explain TEXT NOT NULL,
                desc TEXT NOT NULL,
                reply TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'Pending'
            )
        """)
        
        # Migration: ensure status column exists in existing database
        try:
            cursor.execute("SELECT status FROM doubts LIMIT 1")
        except sqlite3.OperationalError:
            cursor.execute("ALTER TABLE doubts ADD COLUMN status TEXT NOT NULL DEFAULT 'Pending'")
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS chat_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                doubt_id INTEGER NOT NULL,
                sender TEXT NOT NULL,
                message TEXT NOT NULL,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (doubt_id) REFERENCES doubts (id) ON DELETE CASCADE
            )
        """)
        
        self.conn.commit()

        # Check if database is empty to determine if we should migrate or load defaults
        cursor.execute("SELECT COUNT(*) FROM users")
        user_count = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM subjects")
        subject_count = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM doubts")
        doubt_count = cursor.fetchone()[0]

        # If everything is empty, attempt to migrate from Excel or use defaults
        if user_count == 0 and subject_count == 0 and doubt_count == 0:
            excel_exists = os.path.exists(self.users_path) or os.path.exists(self.subjects_path) or os.path.exists(self.doubts_path)
            if excel_exists:
                self._migrate_from_excel()
            else:
                self._load_defaults()
        else:
            # If database has tables populated but might be missing default values, ensure they exist
            self._ensure_defaults()

    def _migrate_from_excel(self):
        try:
            import openpyxl
        except ImportError:
            # If openpyxl is not available, fall back to defaults
            self._load_defaults()
            return

        cursor = self.conn.cursor()
        migrated_any = False

        # Migrate Users
        if os.path.exists(self.users_path):
            try:
                wb = openpyxl.load_workbook(self.users_path)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 3 and row[0]:
                        username = str(row[0]).strip()
                        password = str(row[1]).strip() if row[1] is not None else ""
                        role = str(row[2]).strip() if row[2] is not None else ""
                        cursor.execute(
                            "INSERT OR IGNORE INTO users (username, password, role) VALUES (?, ?, ?)",
                            (username, password, role)
                        )
                wb.close()
                migrated_any = True
            except Exception as e:
                print(f"Error migrating users from excel: {e}")

        # Migrate Subjects
        if os.path.exists(self.subjects_path):
            try:
                wb = openpyxl.load_workbook(self.subjects_path)
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        subject = str(row[0]).strip()
                        cursor.execute(
                            "INSERT OR IGNORE INTO subjects (subject) VALUES (?)",
                            (subject,)
                        )
                wb.close()
                migrated_any = True
            except Exception as e:
                print(f"Error migrating subjects from excel: {e}")

        # Migrate Doubts
        if os.path.exists(self.doubts_path):
            try:
                wb = openpyxl.load_workbook(self.doubts_path)
                sheet = wb.active
                headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
                expected_headers = ["id", "student", "subject", "teacher", "severity", "urgent", "explain", "desc", "reply"]
                
                if headers == expected_headers:
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 9 and row[0] is not None:
                            try:
                                d_id = int(row[0])
                            except (ValueError, TypeError):
                                continue
                            cursor.execute(
                                """INSERT OR IGNORE INTO doubts 
                                (id, student, subject, teacher, severity, urgent, explain, desc, reply) 
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                                (
                                    d_id,
                                    str(row[1]) if row[1] is not None else "",
                                    str(row[2]) if row[2] is not None else "",
                                    str(row[3]) if row[3] is not None else "",
                                    str(row[4]) if row[4] is not None else "Low",
                                    str(row[5]) if row[5] is not None else "No",
                                    str(row[6]) if row[6] is not None else "No",
                                    str(row[7]) if row[7] is not None else "",
                                    str(row[8]) if row[8] is not None else ""
                                )
                            )
                else:
                    col_map = {name: idx for idx, name in enumerate(headers) if name is not None}
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        def get_val(col_name, default=""):
                            if col_name in col_map and col_map[col_name] < len(row):
                                val = row[col_map[col_name]]
                                return val if val is not None else default
                            return default

                        try:
                            d_id_raw = get_val("id", None)
                            if d_id_raw is None:
                                continue
                            d_id = int(d_id_raw)
                        except (ValueError, TypeError):
                            continue

                        cursor.execute(
                            """INSERT OR IGNORE INTO doubts 
                            (id, student, subject, teacher, severity, urgent, explain, desc, reply) 
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            (
                                d_id,
                                get_val("student"),
                                get_val("subject"),
                                get_val("teacher"),
                                get_val("severity", "Low"),
                                get_val("urgent", "No"),
                                get_val("explain", "No"),
                                get_val("desc"),
                                get_val("reply")
                            )
                        )
                wb.close()
                migrated_any = True
            except Exception as e:
                print(f"Error migrating doubts from excel: {e}")

        self.conn.commit()

        if migrated_any:
            # Backup Excel files by renaming them to .bak
            for path in [self.users_path, self.subjects_path, self.doubts_path]:
                if os.path.exists(path):
                    try:
                        os.rename(path, path + ".bak")
                    except Exception as e:
                        print(f"Failed to rename {path} to backup: {e}")

        # Ensure we have defaults if migration was partial
        self._ensure_defaults()

    def _load_defaults(self):
        cursor = self.conn.cursor()
        for uinfo in self.default_users:
            cursor.execute(
                "INSERT OR IGNORE INTO users (username, password, role) VALUES (?, ?, ?)",
                (uinfo.get("username", ""), uinfo.get("password", ""), uinfo.get("role", ""))
            )
        for s in self.default_subjects:
            cursor.execute(
                "INSERT OR IGNORE INTO subjects (subject) VALUES (?)",
                (s,)
            )
        self.conn.commit()

    def _ensure_defaults(self):
        cursor = self.conn.cursor()
        for uinfo in self.default_users:
            cursor.execute(
                "INSERT OR IGNORE INTO users (username, password, role) VALUES (?, ?, ?)",
                (uinfo.get("username", ""), uinfo.get("password", ""), uinfo.get("role", ""))
            )
        for s in self.default_subjects:
            cursor.execute(
                "INSERT OR IGNORE INTO subjects (subject) VALUES (?)",
                (s,)
            )
        self.conn.commit()

    def authenticate(self, username, password):
        """Returns role string ('student' or 'teacher') if valid, else None."""
        u_clean = str(username).strip()
        p_clean = str(password).strip()
        
        cursor = self.conn.cursor()
        cursor.execute("SELECT password, role FROM users WHERE username = ?", (u_clean,))
        row = cursor.fetchone()
        if row and row[0] == p_clean:
            return row[1]
        return None

    def register_user(self, username, password, role):
        """Registers a new user in the database. Returns True if successful, False if username exists."""
        u_clean = str(username).strip()
        p_clean = str(password).strip()
        r_clean = str(role).strip().lower()
        
        if not u_clean or not p_clean or r_clean not in ["student", "teacher"]:
            return False
            
        cursor = self.conn.cursor()
        try:
            cursor.execute(
                "INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                (u_clean, p_clean, r_clean)
            )
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False

    def add_doubt(self, student, subject, teacher, severity, urgent, explain, desc):
        """Adds a new doubt dictionary to the store."""
        cursor = self.conn.cursor()
        cursor.execute(
            """INSERT INTO doubts (student, subject, teacher, severity, urgent, explain, desc, reply, status)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (student, subject, teacher, severity, urgent, explain, desc, "", "Pending")
        )
        self.conn.commit()
        
        doubt_id = cursor.lastrowid
        return {
            "id": doubt_id,
            "student": student,
            "subject": subject,
            "teacher": teacher,
            "severity": severity,
            "urgent": urgent,
            "explain": explain,
            "desc": desc,
            "reply": "",
            "status": "Pending"
        }

    def submit_reply(self, doubt_id, reply_text):
        """Finds doubt by ID and sets the reply text."""
        cursor = self.conn.cursor()
        cursor.execute(
            "UPDATE doubts SET reply = ?, status = CASE WHEN status = 'Pending' THEN 'Replied' ELSE status END WHERE id = ?",
            (reply_text, doubt_id)
        )
        self.conn.commit()
        return cursor.rowcount > 0

    def get_doubts_for_student(self, student_username):
        """Returns list of doubts submitted by a specific student."""
        cursor = self.conn.cursor()
        cursor.execute(
            "SELECT id, student, subject, teacher, severity, urgent, explain, desc, reply, status FROM doubts WHERE student = ?",
            (student_username,)
        )
        rows = cursor.fetchall()
        return [
            {
                "id": r[0],
                "student": r[1],
                "subject": r[2],
                "teacher": r[3],
                "severity": r[4],
                "urgent": r[5],
                "explain": r[6],
                "desc": r[7],
                "reply": r[8],
                "status": r[9]
            }
            for r in rows
        ]

    def get_all_doubts(self):
        """Returns all doubts (for teachers)."""
        cursor = self.conn.cursor()
        cursor.execute(
            "SELECT id, student, subject, teacher, severity, urgent, explain, desc, reply, status FROM doubts"
        )
        rows = cursor.fetchall()
        return [
            {
                "id": r[0],
                "student": r[1],
                "subject": r[2],
                "teacher": r[3],
                "severity": r[4],
                "urgent": r[5],
                "explain": r[6],
                "desc": r[7],
                "reply": r[8],
                "status": r[9]
            }
            for r in rows
        ]

    def get_doubts_for_teacher(self, teacher_username):
        """Returns doubts received by a specific teacher."""
        cursor = self.conn.cursor()
        cursor.execute(
            "SELECT id, student, subject, teacher, severity, urgent, explain, desc, reply, status FROM doubts WHERE teacher = ?",
            (teacher_username,)
        )
        rows = cursor.fetchall()
        return [
            {
                "id": r[0],
                "student": r[1],
                "subject": r[2],
                "teacher": r[3],
                "severity": r[4],
                "urgent": r[5],
                "explain": r[6],
                "desc": r[7],
                "reply": r[8],
                "status": r[9]
            }
            for r in rows
        ]

    def resolve_doubt(self, doubt_id):
        """Sets the doubt status to 'Resolved'."""
        cursor = self.conn.cursor()
        cursor.execute(
            "UPDATE doubts SET status = 'Resolved' WHERE id = ?",
            (doubt_id,)
        )
        self.conn.commit()
        return cursor.rowcount > 0

    def unresolve_doubt(self, doubt_id):
        """Sets the doubt status back to 'Pending'."""
        cursor = self.conn.cursor()
        cursor.execute(
            "UPDATE doubts SET status = 'Pending' WHERE id = ?",
            (doubt_id,)
        )
        self.conn.commit()
        return cursor.rowcount > 0

    def get_subjects(self):
        """Returns list of dynamic subjects."""
        cursor = self.conn.cursor()
        cursor.execute("SELECT subject FROM subjects")
        rows = cursor.fetchall()
        return [r[0] for r in rows]

    def get_teachers(self):
        """Returns list of dynamic teachers."""
        cursor = self.conn.cursor()
        cursor.execute("SELECT username FROM users WHERE role = 'teacher'")
        rows = cursor.fetchall()
        return [r[0] for r in rows]

    def get_chat_messages(self, doubt_id):
        """Returns all chat messages for a given doubt ID."""
        cursor = self.conn.cursor()
        cursor.execute(
            "SELECT sender, message, timestamp FROM chat_messages WHERE doubt_id = ? ORDER BY timestamp ASC",
            (doubt_id,)
        )
        rows = cursor.fetchall()
        return [
            {
                "sender": r[0],
                "message": r[1],
                "timestamp": r[2]
            }
            for r in rows
        ]

    def add_chat_message(self, doubt_id, sender, message):
        """Adds a new chat message to the database."""
        cursor = self.conn.cursor()
        cursor.execute(
            "INSERT INTO chat_messages (doubt_id, sender, message) VALUES (?, ?, ?)",
            (doubt_id, sender, message)
        )
        self.conn.commit()

    def __del__(self):
        # Close connection when object is garbage collected
        try:
            self.conn.close()
        except Exception:
            pass
