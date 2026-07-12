import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
# ═════════════════════════════
# LOAD USERS
# ═════════════════════════════
def load_users():
    if not os.path.exists("users.xlsx"):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["username", "password", "role"])
        sheet.append(["student1", "s123", "student"])
        sheet.append(["teacher1", "t123", "teacher"])
        wb.save("users.xlsx")
    wb = openpyxl.load_workbook("users.xlsx")
    sheet = wb.active

    users = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        username, password, role = row
        users[username] = {"password": password, "role": role}

    return users


# ═════════════════════════════
# LOAD DOUBTS
# ═════════════════════════════
def load_doubts():
    if not os.path.exists("doubts.xlsx"):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["id", "student", "subject", "desc", "reply"])
        wb.save("doubts.xlsx")
    wb = openpyxl.load_workbook("doubts.xlsx")
    sheet = wb.active

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row
        data.append({
            "id": row[0],
            "student": row[1],
            "subject": row[2],
            "desc": row[3],
            "reply": row[4] or ""
        })

    return data


# ═════════════════════════════
# SAVE DOUBTS
# ═════════════════════════════
def save_doubts():
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.append(["id", "student", "subject", "desc", "reply"])

    for d in doubts:
        sheet.append([
            d["id"],
            d["student"],
            d["subject"],
            d["desc"],
            d["reply"]
        ])

    wb.save("doubts.xlsx")


# ═════════════════════════════
# INITIAL LOAD
# ═════════════════════════════
USERS = load_users()
doubts = load_doubts()

if doubts:
    _id = [max(d["id"] for d in doubts) + 1]
else:
    _id = [1]


# ═════════════════════════════
# LOGIN
# ═════════════════════════════
def show_login(root):
    clear(root)

    tk.Label(root, text="Login").pack()

    tk.Label(root, text="Username").pack()
    user_e = tk.Entry(root)
    user_e.pack()

    tk.Label(root, text="Password").pack()
    pass_e = tk.Entry(root, show="*")
    pass_e.pack()
    
    def login(event=None):
        u = user_e.get()
        p = pass_e.get()

        if u in USERS and USERS[u]["password"] == p:
            if USERS[u]["role"] == "student":
                show_student(root, u)
            else:
                show_teacher(root, u)
        else:
            messagebox.showerror("Error", "Invalid login")
    pass_e.bind("<Return>",login)
    tk.Button(root, text="Login", command=login).pack()
    


# ═════════════════════════════
# HEADER
# ═════════════════════════════
def header(root, title):
    tk.Label(root, text=title).pack()
    tk.Button(root, text="Logout", command=lambda: show_login(root)).pack(anchor="ne")


# ═════════════════════════════
# STUDENT DASHBOARD
# ═════════════════════════════
def show_student(root, username):
    clear(root)

    header(root, f"Student: {username}")

    tk.Label(root, text="Subject").pack()
    subj_e = tk.Entry(root)
    subj_e.pack()

    tk.Label(root, text="Description").pack()
    desc_t = tk.Text(root, height=3)
    desc_t.pack()

    def submit():
        s = subj_e.get()
        d = desc_t.get("1.0", "end").strip()

        if not s or not d:
            messagebox.showwarning("Warning", "Fill all fields")
            return

        doubts.append({
            "id": _id[0],
            "student": username,
            "subject": s,
            "desc": d,
            "reply": ""
        })

        _id[0] += 1
        save_doubts()   # ⭐ save to excel

        subj_e.delete(0, "end")
        desc_t.delete("1.0", "end")

        refresh_list()

    tk.Button(root, text="Submit", command=submit).pack()

    # Table
    cols = ("ID", "Subject", "Description", "Reply")
    tree = ttk.Treeview(root, columns=cols, show="headings")

    for c in cols:
        tree.heading(c, text=c)

    tree.pack()

    def refresh_list():
        tree.delete(*tree.get_children())

        for d in doubts:
            if d["student"] == username:
                tree.insert("", "end", values=(
                    d["id"],
                    d["subject"],
                    d["desc"],
                    d["reply"] or "Pending"
                ))

    refresh_list()


# ═════════════════════════════
# TEACHER DASHBOARD
# ═════════════════════════════
def show_teacher(root, username):
    clear(root)

    header(root, f"Teacher: {username}")

    cols = ("ID", "Student", "Subject", "Description", "Reply")
    tree = ttk.Treeview(root, columns=cols, show="headings")

    for c in cols:
        tree.heading(c, text=c)

    tree.pack()

    def load():
        tree.delete(*tree.get_children())

        for d in doubts:
            tree.insert("", "end", iid=str(d["id"]), values=(
                d["id"],
                d["student"],
                d["subject"],
                d["desc"],
                d["reply"] or ""
            ))

    load()

    tk.Label(root, text="Reply").pack()
    reply_e = tk.Entry(root)
    reply_e.pack()

    def send_reply():
        sel = tree.selection()

        if not sel:
            messagebox.showwarning("Warning", "Select a doubt")
            return

        did = int(sel[0])
        txt = reply_e.get()

        if not txt:
            messagebox.showwarning("Warning", "Write reply")
            return

        for d in doubts:
            if d["id"] == did:
                d["reply"] = txt
                break

        save_doubts()   # ⭐ save to excel

        reply_e.delete(0, "end")
        load()

    tk.Button(root, text="Send Reply", command=send_reply).pack()


# ═════════════════════════════
# HELPERS
# ═════════════════════════════
def clear(root):
    for w in root.winfo_children():
        w.destroy()


# ═════════════════════════════
# MAIN
# ═════════════════════════════
root = tk.Tk()
root.title("Doubt System (Excel Based)")
root.geometry("400x400")

show_login(root)

root.mainloop()